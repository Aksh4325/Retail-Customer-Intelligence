"""
Excel Report Generator
Create Excel reports for retail analysis
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelReporter:
    """Generate Excel reports"""
    
    def __init__(self, output_dir='excel'):
        self.output_dir = output_dir
    
    def create_customer_rfm_report(self, rfm_data, filename='customer_rfm.xlsx'):
        """Create customer RFM report"""
        
        print("Creating Customer RFM Report...")
        
        # Select relevant columns
        report_data = rfm_data[[
            'customer_id', 'recency', 'frequency', 'monetary',
            'R_score', 'F_score', 'M_score', 'RFM_score', 
            'segment', 'CLV'
        ]].copy()
        
        # Sort by monetary (highest spenders first)
        report_data = report_data.sort_values('monetary', ascending=False)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Customer RFM Data"
        
        # Title
        ws['A1'] = "Customer RFM Analysis Report"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:J1')
        
        # Write data
        for r in dataframe_to_rows(report_data, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[3]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Format numbers
        for row in range(4, len(report_data) + 4):
            ws.cell(row=row, column=4).number_format = '₹#,##0.00'  # Monetary
            ws.cell(row=row, column=10).number_format = '₹#,##0.00'  # CLV
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 15
        
        filepath = f"{self.output_dir}/{filename}"
        wb.save(filepath)
        print(f"✓ Saved: {filepath}")
        
        return filepath
    
    def create_segment_summary_report(self, segment_summary, filename='segment_summary.xlsx'):
        """Create segment summary report"""
        
        print("Creating Segment Summary Report...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Segment Summary"
        
        # Title
        ws['A1'] = "Customer Segment Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = ['Segment', 'Customers', 'Total Revenue', 'Revenue %', 
                  'Avg Frequency', 'Avg Recency']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row_num = 4
        for _, row in segment_summary.iterrows():
            ws.cell(row=row_num, column=1).value = row['segment']
            ws.cell(row=row_num, column=2).value = row['customer_count']
            ws.cell(row=row_num, column=3).value = row['total_revenue']
            ws.cell(row=row_num, column=4).value = row['revenue_pct']
            ws.cell(row=row_num, column=5).value = round(row['avg_frequency'], 2)
            ws.cell(row=row_num, column=6).value = round(row['avg_recency'], 1)
            
            # Format numbers
            ws.cell(row=row_num, column=3).number_format = '₹#,##0.00'
            ws.cell(row=row_num, column=4).number_format = '0.00'
            ws.cell(row=row_num, column=5).number_format = '0.00'
            ws.cell(row=row_num, column=6).number_format = '0.0'
            
            row_num += 1
        
        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[chr(64+col)].width = 18
        
        filepath = f"{self.output_dir}/{filename}"
        wb.save(filepath)
        print(f"✓ Saved: {filepath}")
        
        return filepath
    
    def create_top_customers_report(self, top_customers, filename='top_customers.xlsx'):
        """Create top customers report"""
        
        print("Creating Top Customers Report...")
        
        # Select top 50
        top_50 = top_customers.head(50)
        
        report_data = top_50[[
            'customer_id', 'monetary', 'frequency', 'recency', 
            'segment', 'CLV'
        ]].copy()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Top Customers"
        
        # Title
        ws['A1'] = "Top 50 Customers by Revenue"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # Write data
        for r in dataframe_to_rows(report_data, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[3]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', fill_type='solid')
        
        # Format numbers
        for row in range(4, len(report_data) + 4):
            ws.cell(row=row, column=2).number_format = '₹#,##0.00'
            ws.cell(row=row, column=6).number_format = '₹#,##0.00'
        
        # Adjust widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        
        filepath = f"{self.output_dir}/{filename}"
        wb.save(filepath)
        print(f"✓ Saved: {filepath}")
        
        return filepath
